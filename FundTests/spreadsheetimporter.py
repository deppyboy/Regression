import openpyxl
import datetime
import pyodbc
from fund import ORACLESTRING

FILENAME = "I:\\Data\\Actuary\\Risk Management - Equity Market\\Fund Analysis\\Return Data\\PLFA\\Individual Fund Returns 1-2-02 to 6-28-13.xlsx"

def importssdata(filenames):
    cnxn = pyodbc.connect(ORACLESTRING)
    c = cnxn.cursor()
    #c.execute('delete from funddata;')
    #c.commit()
    for filename in filenames:
        wb = openpyxl.load_workbook(filename)
        ws = wb.get_sheet_by_name('Fund Returns - Daily')
        colnum = 2
        while not(ws.cell(row=3,column=colnum).value==None):
            dates, returns = [],[]
            rownum = 3
            while ws.cell(row=rownum,column=colnum).value==None:
                rownum+=1
            while not(ws.cell(row=rownum,column=colnum).value==None):
                returnval = ws.cell(row=rownum,column=colnum).value
                if returnval!=0.0:
                    dte = datetime.datetime.strptime(ws.cell(row=rownum,column=1).value,'%m/%d/%Y')
                    dates.append(dte)
                    returns.append(returnval)           
                rownum+=1
            fundname = ws.cell(row=2,column=colnum).value
            print fundname, dates, returns
            c.execute("select * from fundnamecode where mappingname=''" % (fundname))
            row = c.fetchone()
            fundnum = row[2]
            colnum+=1

if __name__=='__main__':
    importssdata(filenames=[FILENAME])